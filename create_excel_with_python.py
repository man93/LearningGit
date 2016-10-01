from openpyxl import Workbook
#workbook object creation
wb = Workbook()
#gives active work sheet
ws = wb.active
#gives title to sheet
ws.title = "Sample_excel.xlsx"
#change default color of title_tab
ws.sheet_properties.tabColor = "1072BA"
source = wb.active
target = wb.copy_worksheet(source)
target.title = "Sample2.xlsx"
#how to put value directly to cell
val1 = ws.cell(row=1, column=2 ,value=10)
val2 = ws.cell(row=2, column=2, value=11)
#using loop,put values in the cells 
for i in range(3,101):
	for j in range(3,101):
		ws.cell(row=i, column=j, value=i)
#saving the workbook
wb.save('Sample_excel.xlsx')

